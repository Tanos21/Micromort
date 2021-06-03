import PySimpleGUI as sg
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
from tensorflow import keras
from tensorflow.keras.layers import Dense,Flatten
import pandas as pd
import numpy as np
import random
import matplotlib.pyplot as plt

    

sg.ChangeLookAndFeel('GreenTan')      

# ------ Menu Definition ------ #      
menu_def = [['File', ['Open', 'Save', 'Exit', 'Properties']],      
            ['Edit', ['Paste', ['Special', 'Normal', ], 'Undo'], ],      
            ['Help', 'About...'], ]      

# ------ Column Definition ------ #      
column1 = [[sg.Text('Column 1', background_color='#F7F3EC', justification='center', size=(10, 1))],      
            [sg.Spin(values=('Spin Box 1', '2', '3'), initial_value='Spin Box 1')],      
            [sg.Spin(values=('Spin Box 1', '2', '3'), initial_value='Spin Box 2')],      
            [sg.Spin(values=('Spin Box 1', '2', '3'), initial_value='Spin Box 3')]]      

layout = [      
    [sg.Menu(menu_def, tearoff=True)],         
    [sg.Text('Введите данные')],
    [sg.Text('Введите свой возраст')],
    [sg.InputText(' ')],
    
        [sg.Frame(layout=[           
    [sg.Radio('Мужской', "RADIO0", default=True, size=(10,1)),
     sg.Radio('Женский', "RADIO0")]], title='Пол',title_color='black', relief=sg.RELIEF_SUNKEN)],
    
    [sg.Frame(layout=[           
    [sg.Radio('нет', "RADIO1", default=True, size=(10,1)),
     sg.Radio('3-5 в день', "RADIO1"),
     sg.Radio('6-10 в день',"RADIO1",size=(10,1)),
     sg.Radio('10-20 в день', "RADIO1")]], title='Курить',title_color='black', relief=sg.RELIEF_SUNKEN)],

        [sg.Frame(layout=[           
    [sg.Radio('нет', "RADIO2", default=True, size=(10,1)),
     sg.Radio('каждый день', "RADIO2"),
     sg.Radio('раз в неделю',"RADIO2", size=(10,1)),
     sg.Radio('раз в месяц', "RADIO2")]], title='Пить',title_color='black', relief=sg.RELIEF_SUNKEN)],

        [sg.Frame(layout=[           
    [sg.Radio('нет', "RADIO3", default=True, size=(10,1)),
     sg.Radio('автобус', "RADIO3"),
     sg.Radio('машина',"RADIO3",  size=(10,1)),
     sg.Radio('мотоцикл', "RADIO3")]], title='Транспорт',title_color='black', relief=sg.RELIEF_SUNKEN)],
    
      [sg.Button('Go')]
]      


window = sg.Window('Micromort', layout, default_element_size=(40, 1), grab_anywhere=False)      
event, values = window.read()

    
print( values[1], values[2])

if values[2]:
    val1=1
else: val1=0

if values[4]:
    val2=0
elif values[5]:
    val2=1
elif values[6]:
    val2=2
else:val2=3

if values[8]:
    val3=0
elif values[9]:
    val3=1
elif values[10]:
    val3=2
else:val3=3

if values[12]:
    val4=0
elif values[13]:
    val4=1
elif values[14]:
    val4=2
else:val4=3

print( values[1],val1, val2,val3,val4)



wb = load_workbook('results.xlsx')
ws = wb.active
ws.title = 'testexel'


ws['A366'] = values[1]
ws['B366'] = val1
ws['C366'] = val2
ws['D366'] = val3
ws['E366'] = val4

wb.save('results.xlsx')



data_frame = pd.read_excel("results.xlsx")
input_names = ["Age","wm","курить","пить","транспорт"]
output_names = ["итог"]





max_age = 100
max_k = 10
encoders = {"Age": lambda age: [age/max_age],
            "wm": lambda wm: [wm],
            "курить": lambda kur:{0:[0,0,0],1:[0,0,1],2:[0,1,0],3:[1,0,0]}.get(kur),
            "пить": lambda pit: {0:[0,0,0],1:[0,0,1],2:[0,1,0],3:[1,0,0]}.get(pit),
            "транспорт": lambda tr:{0:[0,0,0],1:[0,0,1],2:[0,1,0],3:[1,0,0]}.get(tr),
            "итог": lambda kk:[kk]
            }



def detaframe_to_dict(df):
    result=dict()
    for column in df.columns:
        values = data_frame[column].values
        result[column]=values
    return result


def make_supervised(df):
    raw_input_data=data_frame[input_names]
    raw_output_dat=data_frame[output_names]
    return{"inputs":detaframe_to_dict(raw_input_data),
           "outputs":detaframe_to_dict(raw_output_dat)}


def encode(data):
    vectors = []
    for data_name, data_values in data.items():
        encoded = list(map(encoders[data_name],data_values))
        vectors.append(encoded)
    #print(vectors)
    formatted=[]
    for vector_raw in list(zip(*vectors)):
        vector=[]
        for element in vector_raw:
            for e in element:
                vector.append(e)
        formatted.append(vector)
    return formatted
supervised = make_supervised(data_frame)
encoded_inputs = np.array(encode(supervised["inputs"]))
encoded_outputs = np.array(encode(supervised["outputs"]))


#print(encoded_outputs)

train_x = encoded_inputs[:364]
train_y = encoded_outputs[:364]

test_x = encoded_inputs[364:365]

print(test_x)

model = keras.Sequential()
model.add(keras.layers.Dense(30, input_dim=11, activation = "relu"))
model.add(keras.layers.Dense(units = 1,input_dim=30, activation = "linear"))

keras.optimizers.SGD(learning_rate=0.01, momentum=0.4, nesterov=True)
model.compile(loss="mse", optimizer="sgd", metrics=["accuracy"])
fit_result=model.fit(x=train_x, y=train_y, epochs=400, validation_split=0.3)

plt.title("losses train\validation")
plt.plot(fit_result.history["loss"],label="train")
plt.plot(fit_result.history["val_loss"],label="Validation")
plt.legend()
plt.show()

plt.title("Accuracies\validation")
plt.plot(fit_result.history["accuracy"],label="train")
plt.plot(fit_result.history["val_accuracy"],label="Validation")
plt.legend()
plt.show()

predicted_test = model.predict(test_x)
real_data = data_frame.iloc[364:365][input_names+output_names]
real_data["Itog"]=predicted_test
print(real_data)






Ag=int(values[1])
ver=1000000-int(predicted_test)
ver1=ver/1000000
ver2=ver1**(365*Ag)
ver3=ver2*100
ver4=100-ver3
print(ver,ver1,ver2,ver3,ver4)

if(int(values[1])<=54): q =int(predicted_test+1.2)
else: q=int(predicted_test*2-4)
Ag1=int(values[1])+5
ver0=1000000-int(q)
ver11=ver0/1000000
ver21=ver11**(365*Ag1)
ver31=ver21*100
ver41=100-ver31

Ag2=int(values[1])+7
ver01=1000000-int(q)
ver111=ver01/1000000
ver211=ver111**(365*Ag2)
ver311=ver211*100
ver411=100-ver311

if (int(val2)==1): VR=float(predicted_test)-4
elif (int(val2)==2): VR=float(predicted_test)-5
elif (int(val2)==3): VR=float(predicted_test)-10
else: VR=""

if (int(val2)!=0): rec="Бросить курить. Курение повышает риск смерти. Если вы сейчас бросите купить, то Ваша еденица риска будет равна:"
else:rec=""
if (int(val4) == 3): rec1="Смените вид транспорта"
else:rec1=""



sg.popup('Результат',
         "Еденица риска равна:",
             predicted_test,
         "Вероятность смерти в этом возрасте равна:", round(ver4,1),
         "Вероятность смерти через 5 лет равна:", round(ver41,1),
         "Вероятность смерти через 7 лет равна:", round(ver411,1),
         "Рекомендация: ", rec,round(VR,3),rec1
         )

window.close()
